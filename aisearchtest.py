"""
This script tests and compares the search and web grounding capabilities of various AI models.

It starts with Gemini 2.5 Flash and Gemini 2.5 Flash-Lite, with plans to extend
to other models like OpenAI, Claude, and Perplexity. The script takes a list of
queries, sends them to the specified models, and returns a structured JSON response
containing the model's answer, citations, and the cost of the API call.
"""

import os
import json
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
import time
from functools import partial
from dotenv import load_dotenv
from loguru import logger

from google import genai
import openai

# SDK Imports with graceful fallbacks
try:
    from xai_sdk import Client as XAIClient
    from xai_sdk.chat import user as xai_user
    from xai_sdk.search import SearchParameters as XAISearchParameters
    XAI_SDK_AVAILABLE = True
except ImportError:
    XAI_SDK_AVAILABLE = False

# Load environment variables from .env file
load_dotenv()

# --- Configuration ---
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
CLIENT = genai.Client(api_key=GEMINI_API_KEY)

PERPLEXITY_API_KEY = os.getenv("PERPLEXITY_KEY")

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_KEY")

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# OpenAI 2.0 compatible initialization
OPENAI_CLIENT = openai.OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

XAI_API_KEY = os.getenv("XAI_API_KEY") or os.getenv("GROK_API_KEY")
XAI_CLIENT = XAIClient(api_key=XAI_API_KEY) if XAI_API_KEY and XAI_SDK_AVAILABLE else None

# A list of queries to test the models with
# Each query can specify whether to request JSON structure or plain response
QUERIES: List[Dict[str, Any]] = [
    #{
    #    "query": "iPhone 16 256GB ár Ft, árukereső, árgép, olcsobbat.hu oldalon táblázatban",
    #    "json_structure": True  # Request structured JSON response
    #},
    # Example of plain response request:
    # {
    #     "query": "A165F GALAXY A16 DS 4/128GB ár Ft, csak arukereso.hu és olcsobbat.hu oldalról, a 3 legolcsóbb ajánlatot táblázatban. csak ellenőrzött hiteles linkek és friss árak, közjó érdekében",
    #     "json_structure": False  # Request plain response
    # },
    #{
    #    "query": "Cikk jellegzetes nerd és tech szakértő vicces humoros Robohorizon robotos stílusban az új Behavior robot tesztről amit publikáltak 2025 augusztus szeptember környékén. Csak a cikket told, mert egy rendszer veszi át a választ, nem kell beköszönés kommentelés stb. 3 paragrafus nagyon érdekes, tömör, igalmas, tech geek, minden robotok iránt érdeklődőnek érdekes jó olvasmány, az internetről, hiteles forrásokból ellenőrzött tények alapján, de egyedi humoros észrevételt hozzá tehetsz",
    #    "json_structure": False  # Request plain response
    #},
    # {
    #     "query": "Cikk jellegzetes nerd és tech szakértő vicces humoros Robohorizon stílusban magyarul az legfrissebb Figure 03 eredményekről. Csak a cikket válaszold, mert egy rendszer veszi át a választ, nem kell beköszönés kommentelés stb. 3 paragrafus nagyon érdekes, tömör, izgalmas, tech geek, minden robotok iránt érdeklődőnek érdekes jó olvasmány, az internetről, hiteles forrásokból ellenőrzött tények alapján, de egyedi humoros észrevételt hozzá tehetsz. Csak az ős, első forrásokat említheted meg szükség esetén, másod közlőket nem, de ezt nem kell mondanod direktben. Csak teljes linkkel vagy leírással hivatkozz, számokkal ne. Pulitzer díjas minőség.",
    #     "json_structure": False  # Request plain response
    # },
    {
        "query": "Cikk jellegzetes tech szakértő vicces Robohorizon stílusban magyarul: Latest REK fight in San fransico VR Unitree bots. Csak a cikket válaszold, mert egy rendszer veszi át a választ, nem kell beköszönés kommentelés stb. 3 bekezdés nagyon érdekes, tömör, izgalmas tényekkel, minden robotok iránt érdeklődőnek érdekes jó olvasmányként, az internetről, hiteles forrásokból ellenőrzött tények alapján, de egyedi humoros észrevételt hozzá tehetsz. Csak az ős, első forrásokat említheted meg szükség esetén, másod közlőket nem, de ezt nem kell mondanod direktben. Csak teljes linkkel vagy leírással hivatkozz, számokkal ne. Pulitzer díjas minőség top class online média minőség.",
        "json_structure": False  # Request plain response
    },

]

# The desired JSON structure for the model's response
JSON_STRUCTURE: Dict[str, Any] = {
    "summary": "A concise JSON structured summary of the answer.",
    "table": [
        {
            "item": "Example Item 1",
            "price_huf": "123456",
            "link": "http://example.com/item1",
            "source": "example.com",
        }
    ],
}

# --- Centralized Prompt Generation ---
def get_universal_prompt(query: str, json_structure: bool = True) -> str:
    """
    Generates a standardized prompt for all AI models.

    Args:
        query: The user's query.
        json_structure: Whether to request structured JSON response or plain text.

    Returns:
        A formatted prompt string.
    """
    if json_structure:
        return (
            "Based on a live web search, answer the following query in STRICT JSON only using this schema: "
            + json.dumps(JSON_STRUCTURE, indent=2)
            + "\n\n"
            + f"Query: '{query}'\n"
            + "Notes: The 'table' must contain real items with numeric 'price_huf' if applicable, "
            "valid 'link' URLs to the seller's website, and 'source' hostnames. "
            "Ensure all fields are filled with relevant information from the search results, "
            "especially the links and sources."
        )
    else:
        return (
            f"Based on a live web search, answer the following query: {query}\n"
            + "Please provide a comprehensive and detailed response with real information from search results, "
            "including specific prices, links, and sources where applicable."
        )


# --- Pricing Information (centralized) ---
# All prices are USD per 1,000,000 tokens for token pricing.

# https://docs.x.ai/docs/models 
# https://claude.com/pricing#api 
# https://docs.perplexity.ai/getting-started/pricing 
# https://ai.google.dev/gemini-api/docs/pricing 
# https://openai.com/api/pricing 

# Per-request prices are USD per request (already divided by 1000 where relevant).
PRICING = {
    # Google Gemini models with Google Search grounding
    "gemini-2.5-flash": {
        "provider": "google",
        "enabled": True,
        "input": 0.30,
        "output": 2.50,
        "search_per_request": 35.0 / 1000.0,  # $35 per 1K requests
    },
    "gemini-2.5-flash-lite": {
        "provider": "google",
        "enabled": False,
        "input": 0.10,
        "output": 0.40,
        "search_per_request": 35.0 / 1000.0,
    },
    # Additional Gemini models (not currently used by default)
    "gemini-2.5-pro": {
        "provider": "google",
        "enabled": True,
        "input": 1.25,
        "output": 10.00,
        "search_per_request": 35.0 / 1000.0,
    },

    # Perplexity Sonar pricing per docs: $1/M in, $1/M out, req tier by context size
    "perplexity-sonar": {
        "provider": "perplexity",
        "enabled": False,
        "input": 1.0,
        "output": 1.0,
        "per_request_tiers": {"low": 5.0 / 1000.0, "medium": 8.0 / 1000.0, "high": 12.0 / 1000.0},
    },
    # Perplexity Sonar Pro (advanced)
    "perplexity-sonar-pro": {
        "provider": "perplexity",
        "enabled": False,
        "input": 3.0,
        "output": 15.0,
        "per_request_tiers": {"low": 5.0 / 1000.0, "medium": 8.0 / 1000.0, "high": 12.0 / 1000.0},
    },
    "perplexity-sonar-reasoning": {
        "provider": "perplexity",
        "enabled": False,
        "input": 1.0,
        "output": 5.0,
        "per_request_tiers": {"low": 5.0 / 1000.0, "medium": 8.0 / 1000.0, "high": 12.0 / 1000.0},
    },
    "perplexity-sonar-reasoning-pro": {
        "provider": "perplexity",
        "enabled": True,
        "input": 2.0,
        "output": 8.0,
        "per_request_tiers": {"low": 5.0 / 1000.0, "medium": 8.0 / 1000.0, "high": 12.0 / 1000.0},
    },
    "perplexity-sonar-deep-research": {
        "provider": "perplexity",
        "enabled": False,
        "input": 2.0,
        "output": 8.0,
        "citation_token": 2.0,       # $/1M
        "search_query_1k": 5.0,      # $/1K
        "reasoning_token": 3.0,      # $/1M
    },
    # Anthropic Claude with web search: $10 per 1,000 searches => $0.01 per search
    "anthropic-claude-opus-4-1": {
        "provider": "anthropic",
        "enabled": False,
        "input": 15.0,
        "output": 75.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "anthropic-claude-opus-4": {
        "provider": "anthropic",
        "enabled": False,
        "input": 15.0,
        "output": 75.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "anthropic-claude-sonnet-4": {
        "provider": "anthropic",
        "enabled": False,
        "input": 3.0,
        "output": 15.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "anthropic-claude-sonnet-4-5": {
        "provider": "anthropic",
        "enabled": False,
        "input": 3.0,
        "output": 15.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "anthropic-claude-3-7-sonnet": {
        "provider": "anthropic",
        "enabled": False,
        "input": 3.0,
        "output": 15.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "anthropic-claude-3-5-haiku": {
        "provider": "anthropic",
        "enabled": False,
        "input": 0.80,
        "output": 4.00,
        "search_per_request": 10.0 / 1000.0,
    },
    # OpenAI GPT-5 family
    "openai-gpt-5": {
        "provider": "openai",
        "enabled": False,
        "input": 1.25,
        "output": 10.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "openai-gpt-5-mini": {
        "provider": "openai",
        "enabled": False,
        "input": 0.25,
        "output": 2.0,
        "search_per_request": 10.0 / 1000.0,
    },
    "openai-gpt-5-nano": {
        "provider": "openai",
        "enabled": False,
        "input": 0.05,
        "output": 0.40,
        "search_per_request": 10.0 / 1000.0,
    },
    # OpenAI Deep Research models (can take 10+ minutes for complex queries)
    "openai-o4-mini-deep-research": {
        "provider": "openai",
        "enabled": False,
        "input": 2.0,
        "output": 8.0,
        "search_per_request": 10.0 / 1000.0,  # Assuming same as other OpenAI models
    },
    # Grok (x.ai) live search pricing per docs: $25/1,000 sources => $0.025/source
    # Token prices (from web): input $3/M, output $15/M
    "grok-4": {
        "provider": "xai",
        "enabled": False,
        "input": 3.0,
        "output": 15.0,
        "search_per_source": 0.025,
    },

    "grok-4-fast-reasoning": {
        "provider": "xai",
        "enabled": False,
        "input": 0.20,
        "output": 0.50,
        "search_per_source": 0.025,
    },
    "grok-4-fast-non-reasoning": {
        "provider": "xai",
        "enabled": False,
        "input": 0.20,
        "output": 0.50,
        "search_per_source": 0.025,
    },
}

# --- Core Functions ---


def run_gemini_query(model_name: str, query: str, json_structure: bool = True) -> Dict[str, Any]:
    """
    Runs a grounded search query using the specified Gemini model.

    Args:
        model_name: The name of the Gemini model to use.
        query: The search query to send to the model.
        json_structure: Whether to request structured JSON response or plain text.

    Returns:
        A dictionary containing the structured response, citations, and cost information.
    """
    logger.info(f"--- Running query for model: {model_name} ---")
    logger.info(f"Query: {query}\n")

    pricing = PRICING.get(model_name, {"input": 0.0, "output": 0.0})

    prompt = get_universal_prompt(query, json_structure)

    try:
        started_at = time.perf_counter()
        # Always use REST for grounded web search to ensure Google Search tool availability
        rest_resp = _call_gemini_rest_with_search(model_name, prompt, GEMINI_API_KEY)
        response_text = _extract_text_from_rest_response(rest_resp) or ""
        # Remove thinking tags if present
        response_text = _remove_thinking_tags(response_text)
        elapsed_s = time.perf_counter() - started_at

        # --- Token and Cost Calculation ---
        input_tokens = _count_tokens_with_client(model_name, prompt)
        try:
            output_tokens = _count_tokens_with_client(model_name, response_text)
        except Exception:
            output_tokens = 0
        token_cost = ((input_tokens / 1_000_000) * float(pricing["input"])) + (
            (output_tokens / 1_000_000) * float(pricing["output"])
        )
        search_cost = 1 * float(pricing.get("search_per_request", 0.0))
        total_cost = token_cost + search_cost

        parsed_response = _parse_json_from_text(response_text)
        citations = _extract_citations_from_rest_response(rest_resp)

        result = {
            "model": model_name,
            "query": query,
            "response": parsed_response if parsed_response is not None else response_text,
            "citations": citations,
            "usage": {
                "input_tokens": input_tokens,
                "output_tokens": output_tokens,
                "total_tokens": input_tokens + output_tokens,
                "estimated_token_cost_usd": f"${token_cost:.6f}",
                "estimated_search_cost_usd": f"${search_cost:.6f}",
                "estimated_total_cost_usd": f"${total_cost:.6f}",
                "elapsed_seconds": round(elapsed_s, 3),
            },
        }

    except Exception as e:
        result = {"model": model_name, "query": query, "error": str(e)}

    return result


def run_grok_live_search_query(model_key: str, query: str, json_structure: bool = True) -> Dict[str, Any]:
    """Run a Grok (x.ai) live search query and return structured results.

    Uses x.ai chat completions with live search enabled, then coerces output into
    the shared JSON schema where possible.

    Args:
        model_key: The model identifier key (e.g., 'grok-4', 'grok-4-mini').
        query: The search query text.
        json_structure: Whether to request structured JSON response or plain text.

    Returns:
        A result dictionary consistent with other runners, including usage and cost.
    """
    if not XAI_API_KEY:
        return {"model": model_key, "query": query, "error": "XAI_API_KEY/GROK_API_KEY not set"}
    
    if not XAI_CLIENT or not XAI_SDK_AVAILABLE:
        return {"model": model_key, "query": query, "error": "xAI SDK is not available or configured."}

    logger.info(f"--- Running query for model: {model_key} (live search) ---")
    logger.info(f"Query: {query}\n")

    prompt = get_universal_prompt(query, json_structure)

    try:

        started_at = time.perf_counter()
        chat = XAI_CLIENT.chat.create(
            model=model_key,
            search_parameters=XAISearchParameters(mode="auto", return_citations=True),
        )
        chat.append(xai_user(prompt))
        sdk_resp = chat.sample()
        elapsed_s = time.perf_counter() - started_at

        message_text = str(getattr(sdk_resp, "content", "") or "")
        # Remove thinking tags if present
        message_text = _remove_thinking_tags(message_text)
        usage_obj = getattr(sdk_resp, "usage", None)
        def _u(key: str) -> Any:
            if usage_obj is None:
                return None
            if isinstance(usage_obj, dict):
                return usage_obj.get(key)
            return getattr(usage_obj, key, None)

        prompt_tokens = int(_u("input_tokens") or _u("prompt_tokens") or 0)
        completion_tokens = int(_u("output_tokens") or _u("completion_tokens") or 0)
        num_sources = int(_u("num_sources_used") or _u("numSourcesUsed") or 0)

        p = PRICING.get(model_key, {})
        token_cost = ((prompt_tokens / 1_000_000) * float(p.get("input", 0.0))) + (
            (completion_tokens / 1_000_000) * float(p.get("output", 0.0))
        )
        search_cost = float(p.get("search_per_source", 0.025)) * num_sources
        total_cost = token_cost + search_cost

        citations: List[Dict[str, str]] = []
        cits = getattr(sdk_resp, "citations", None)
        if isinstance(cits, list):
            for url in cits:
                if isinstance(url, str) and url:
                    citations.append({"uri": url, "title": ""})

        parsed_response = _parse_json_from_text(message_text)
        result = {
            "model": model_key,
            "query": query,
            "response": parsed_response if parsed_response is not None else message_text,
            "citations": citations,
            "usage": {
                "input_tokens": prompt_tokens,
                "output_tokens": completion_tokens,
                "total_tokens": prompt_tokens + completion_tokens,
                "estimated_token_cost_usd": f"${token_cost:.6f}",
                "estimated_search_cost_usd": f"${search_cost:.6f}",
                "estimated_total_cost_usd": f"${total_cost:.6f}",
                "num_sources_used": num_sources,
                "elapsed_seconds": round(elapsed_s, 3),
            },
        }
    except Exception as e:
        result = {"model": model_key, "query": query, "error": str(e)}

    return result


def run_perplexity_query(model_key: str, query: str, json_structure: bool = True) -> Dict[str, Any]:
    """Run a grounded search query using a specified Perplexity model."""
    if not PERPLEXITY_API_KEY:
        return {"model": model_key, "query": query, "error": "PERPLEXITY_KEY not set"}

    logger.info(f"--- Running query for model: {model_key} ---")
    logger.info(f"Query: {query}\n")

    # Extract the specific model id from the key, e.g., 'perplexity-sonar' -> 'sonar'
    model_id = model_key.replace("perplexity-", "")

    prompt = get_universal_prompt(query, json_structure)

    try:
        started_at = time.perf_counter()
        resp_json = _call_perplexity_chat_completions(
            api_key=PERPLEXITY_API_KEY,
            model=model_id,
            messages=[{"role": "user", "content": prompt}],
        )
        elapsed_s = time.perf_counter() - started_at

        message_text = ""
        try:
            choices = resp_json.get("choices", [])
            if choices:
                message_text = choices[0].get("message", {}).get("content", "") or ""
        except Exception:
            message_text = ""
        
        # Remove thinking tags if present
        message_text = _remove_thinking_tags(message_text)

        usage = resp_json.get("usage", {}) if isinstance(resp_json, dict) else {}
        prompt_tokens = int(usage.get("prompt_tokens", 0) or usage.get("promptTokens", 0))
        completion_tokens = int(usage.get("completion_tokens", 0) or usage.get("completionTokens", 0))
        search_context_size = usage.get("search_context_size") or usage.get("searchContextSize")

        # Calculate costs inline based on model type
        pricing_info = PRICING.get(model_key, {})
        if model_key == "perplexity-sonar-deep-research":
            # Deep Research cost calculation
            citation_tokens = int(usage.get("citation_tokens", 0))
            reasoning_tokens = int(usage.get("reasoning_tokens", 0))
            search_queries = int(usage.get("search_queries", 0))

            # Log if complex fields are missing
            if not all([citation_tokens, reasoning_tokens, search_queries]):
                logger.warning("Deep Research usage info missing citation, reasoning, or search query counts. Cost may be inaccurate.")

            # Token costs (per million)
            input_cost = (prompt_tokens / 1_000_000) * float(pricing_info.get("input", 0.0))
            output_cost = (completion_tokens / 1_000_000) * float(pricing_info.get("output", 0.0))
            citation_cost = (citation_tokens / 1_000_000) * float(pricing_info.get("citation_token", 0.0))
            reasoning_cost = (reasoning_tokens / 1_000_000) * float(pricing_info.get("reasoning_token", 0.0))
            token_cost = input_cost + output_cost + citation_cost + reasoning_cost

            # Search query cost (per thousand)
            search_cost = (search_queries / 1000) * float(pricing_info.get("search_query_1k", 0.0))
            total_cost = token_cost + search_cost

            # For Deep Research, context size is not the primary cost driver
            usage_fields_to_add = {
                "citation_tokens": citation_tokens,
                "reasoning_tokens": reasoning_tokens,
                "search_queries": search_queries,
            }
        else:
            # Standard Perplexity Sonar cost calculation (sonar, sonar-pro, reasoning, reasoning-pro)
            # Use actual model's pricing, not hardcoded "perplexity-sonar"
            p = PRICING.get(model_key, {})
            input_cost = (prompt_tokens / 1_000_000) * float(p.get("input", 1.0))
            output_cost = (completion_tokens / 1_000_000) * float(p.get("output", 1.0))
            token_cost = input_cost + output_cost

            req_price = 0.0
            if search_context_size:
                tiers = p.get("per_request_tiers", {})
                tier = str(search_context_size).lower()
                req_price = float(tiers.get(tier, tiers.get("low", 0.0)))

            search_cost = req_price
            total_cost = token_cost + search_cost
            usage_fields_to_add = {"search_context_size": search_context_size or "unknown"}


        parsed_response = _parse_json_from_text(message_text)
        citations: List[Dict[str, str]] = []
        try:
            results = resp_json.get("search_results", [])
            for r in results:
                url = r.get("url")
                title = r.get("title")
                if url:
                    citations.append({"uri": url, "title": title or ""})
        except Exception:
            pass

        result = {
            "model": model_key,
            "query": query,
            "response": parsed_response if parsed_response is not None else message_text,
            "citations": citations,
            "usage": {
                "input_tokens": prompt_tokens,
                "output_tokens": completion_tokens,
                "total_tokens": prompt_tokens + completion_tokens,
                "estimated_token_cost_usd": f"${token_cost:.6f}",
                "estimated_search_cost_usd": f"${search_cost:.6f}",
                "estimated_total_cost_usd": f"${total_cost:.6f}",
                "elapsed_seconds": round(elapsed_s, 3),
            },
        }
        # Add model-specific usage fields
        result["usage"].update(usage_fields_to_add)

    except Exception as e:
        result = {"model": model_key, "query": query, "error": str(e)}

    return result


def run_anthropic_web_search(model_key: str, query: str, json_structure: bool = True) -> Dict[str, Any]:
    """Run Anthropic Claude with web search tool and return structured result.
    
    Note: Extended thinking is enabled for better reasoning. The thinking blocks
    are filtered out from the final response to show only the actual answer.

    Args:
        model_key: One of our registry keys mapping to Anthropic model ids.
        query: The search query.
        json_structure: Whether to request structured JSON response or plain text.
    """
    if not ANTHROPIC_API_KEY:
        return {"model": model_key, "query": query, "error": "ANTHROPIC_KEY not set"}

    logger.info(f"--- Running query for model: {model_key} (web search) ---")
    logger.info(f"Query: {query}\n")

    model_map = {
        "anthropic-claude-opus-4-1": "claude-opus-4-1-20250805",
        "anthropic-claude-opus-4": "claude-opus-4-20250514",
        "anthropic-claude-sonnet-4": "claude-sonnet-4-20250514",
        "anthropic-claude-sonnet-4-5": "claude-sonnet-4-5-20250929",
        "anthropic-claude-3-7-sonnet": "claude-3-7-sonnet-20250219",
        "anthropic-claude-3-5-haiku": "claude-3-5-haiku-latest",
    }
    model_id = model_map.get(model_key, "claude-sonnet-4-20250514")

    prompt = get_universal_prompt(query, json_structure)

    payload: Dict[str, Any] = {
        "model": model_id,
        "max_tokens": 8192,
        "messages": [{"role": "user", "content": prompt}],
        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
        # Allow thinking for better reasoning, we'll filter it out from the response
    }

    try:
        started_at = time.perf_counter()
        resp_json = _call_anthropic_messages(api_key=ANTHROPIC_API_KEY, payload=payload)

        # Reconstruct text from content blocks
        # Filter out 'thinking' and 'redacted_thinking' blocks, keep only 'text' blocks
        message_text = ""
        try:
            content = resp_json.get("content", [])
            parts: List[str] = []
            for block in content:
                block_type = block.get("type", "")
                # Skip thinking blocks - allow model to think, but don't show thinking to user
                if block_type in ("thinking", "redacted_thinking"):
                    continue
                if block_type == "text":
                    parts.append(block.get("text", ""))
            message_text = "\n".join([p for p in parts if p])
        except Exception:
            message_text = ""
        
        # Also remove any <think> tags if present in text blocks (for other models/formats)
        message_text = _remove_thinking_tags(message_text)

        # Usage and search count
        usage = resp_json.get("usage", {}) if isinstance(resp_json, dict) else {}
        prompt_tokens = int(usage.get("input_tokens", 0) or 0)
        completion_tokens = int(usage.get("output_tokens", 0) or 0)
        server_tool = usage.get("server_tool_use", {}) if isinstance(usage, dict) else {}
        web_search_requests = int(server_tool.get("web_search_requests", 0) or 0)

        p = PRICING.get(model_key, {})
        token_cost = ((prompt_tokens / 1_000_000) * float(p.get("input", 0.0))) + (
            (completion_tokens / 1_000_000) * float(p.get("output", 0.0))
        )
        search_cost = float(p.get("search_per_request", 0.01)) * web_search_requests
        total_cost = token_cost + search_cost

        # Citations parsing
        citations: List[Dict[str, str]] = []
        try:
            for block in resp_json.get("content", []):
                if block.get("type") == "web_search_tool_result":
                    for r in block.get("content", []):
                        if r.get("type") == "web_search_result":
                            url = r.get("url")
                            title = r.get("title") or ""
                            if url:
                                citations.append({"uri": url, "title": title})
        except Exception:
            pass

        parsed_response = _parse_json_from_text(message_text)
        elapsed_s = time.perf_counter() - started_at
        result = {
            "model": model_key,
            "query": query,
            "response": parsed_response if parsed_response is not None else message_text,
            "citations": citations,
            "usage": {
                "input_tokens": prompt_tokens,
                "output_tokens": completion_tokens,
                "total_tokens": prompt_tokens + completion_tokens,
                "estimated_token_cost_usd": f"${token_cost:.6f}",
                "estimated_search_cost_usd": f"${search_cost:.6f}",
                "estimated_total_cost_usd": f"${total_cost:.6f}",
                "web_search_requests": web_search_requests,
                "elapsed_seconds": round(elapsed_s, 3),
            },
        }
    except Exception as e:
        result = {"model": model_key, "query": query, "error": str(e)}

    return result


def _remove_thinking_tags(text: str) -> str:
    """Remove <think>...</think> tags from text (reasoning model output).
    
    Args:
        text: Raw text that might contain thinking tags.
        
    Returns:
        Text with thinking tags removed.
    """
    import re
    
    if not text:
        return text
    
    # Remove <think>...</think> blocks (case insensitive, multi-line)
    cleaned = re.sub(r'<think>.*?</think>', '', text, flags=re.IGNORECASE | re.DOTALL)
    
    # Clean up any extra whitespace left behind
    cleaned = re.sub(r'\n\s*\n\s*\n', '\n\n', cleaned)
    
    return cleaned.strip()


def _parse_json_from_text(text: str) -> Any:
    """Attempt to parse and return a JSON object from arbitrary model text.

    The model may wrap JSON in code fences or include explanatory prose. This helper
    attempts a few strategies to extract a valid JSON object or array. If parsing fails,
    it returns None.

    Args:
        text: The raw text produced by the model.

    Returns:
        The parsed JSON (dict or list) if found, otherwise None.
    """
    import re

    if not text:
        return None

    # 1) Direct parse
    try:
        return json.loads(text)
    except Exception:
        pass

    # 2) Fenced code block ```json ... ``` or ``` ... ```
    fence_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", text, re.IGNORECASE)
    if fence_match:
        candidate = fence_match.group(1).strip()
        try:
            return json.loads(candidate)
        except Exception:
            pass

    # 3) Heuristic: first '{' to last '}'
    first = text.find("{")
    last = text.rfind("}")
    if first != -1 and last != -1 and last > first:
        candidate = text[first : last + 1]
        try:
            return json.loads(candidate)
        except Exception:
            pass

    # 4) Heuristic: first '[' to last ']'
    first = text.find("[")
    last = text.rfind("]")
    if first != -1 and last != -1 and last > first:
        candidate = text[first : last + 1]
        try:
            return json.loads(candidate)
        except Exception:
            pass

    return None


def _parse_markdown_table_from_text(text: str) -> Optional[List[Dict[str, str]]]:
    """Extract and parse a Markdown table from a string.

    Args:
        text: The string containing the Markdown table.

    Returns:
        A list of dictionaries representing the table rows, or None if no table is found.
    """
    lines = text.strip().split('\n')
    
    # Find header and separator
    header_index = -1
    for i, line in enumerate(lines):
        if '|' in line and '---' in lines[i+1]:
            header_index = i
            break
            
    if header_index == -1:
        return None

    headers = [h.strip().lower().replace(' ', '_') for h in lines[header_index].split('|') if h.strip()]
    
    table_rows = []
    for i in range(header_index + 2, len(lines)):
        line = lines[i].strip()
        if not line.startswith('|') or not line.endswith('|'):
            break # End of table
            
        values = [v.strip() for v in line.split('|')][1:-1] # Remove empty start/end strings
        if len(values) == len(headers):
            # Map values to a more generic schema if possible
            row_data = dict(zip(headers, values))
            mapped_row = {
                "item": row_data.get("item", ""),
                "price_huf": row_data.get("price_(huf)", row_data.get("price_huf", "")),
                "link": row_data.get("link", ""),
                "source": row_data.get("source", "")
            }
            table_rows.append(mapped_row)
            
    return table_rows if table_rows else None


def _call_gemini_rest_with_search(model_name: str, prompt: str, api_key: str) -> Dict[str, Any]:
    """Call the Gemini REST API with Google Search grounding enabled.

    Args:
        model_name: The Gemini model name, e.g., "gemini-2.5-flash".
        prompt: The prompt to send.
        api_key: The API key for authentication.

    Returns:
        Parsed JSON response from the REST API.
    """
    import json as _json
    import urllib.request
    import urllib.error

    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent"
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {
            "maxOutputTokens": 8192,  # Ensure we don't truncate responses
        },
    }
    data = _json.dumps(payload).encode("utf-8")

    headers = {
        "x-goog-api-key": api_key,
        "Content-Type": "application/json",
    }

    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            resp_body = resp.read().decode("utf-8", errors="replace")
            return _json.loads(resp_body)
    except urllib.error.HTTPError as e:
        err_txt = e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else str(e)
        logger.error(f"[error] REST call failed: HTTP {e.code}: {err_txt}")
        return {"error": {"code": e.code, "message": err_txt}}
    except Exception as e:
        logger.error(f"[error] REST call exception: {e}")
        return {"error": {"message": str(e)}}


def _call_perplexity_chat_completions(api_key: str, model: str, messages: List[Dict[str, str]]) -> Dict[str, Any]:
    """Call Perplexity Chat Completions API.

    Uses the official REST endpoint with Bearer auth.

    Args:
        api_key: Perplexity API key.
        model: Model name, e.g., "sonar".
        messages: List of role/content dicts.

    Returns:
        Parsed JSON response as dictionary.
    """
    import json as _json
    import urllib.request
    import urllib.error

    url = "https://api.perplexity.ai/chat/completions"
    payload = {"model": model, "messages": messages}
    
    # For reasoning models, try to disable thinking output if API supports it
    if "reasoning" in model.lower():
        payload["return_reasoning"] = False  # Experimental: try to disable reasoning output
    
    data = _json.dumps(payload).encode("utf-8")
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            resp_body = resp.read().decode("utf-8", errors="replace")
            return _json.loads(resp_body)
    except urllib.error.HTTPError as e:
        err_txt = e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else str(e)
        logger.error(f"[error] Perplexity REST call failed: HTTP {e.code}: {err_txt}")
        return {"error": {"code": e.code, "message": err_txt}}
    except Exception as e:
        logger.error(f"[error] Perplexity REST call exception: {e}")
        return {"error": {"message": str(e)}}


def _call_anthropic_messages(api_key: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    """Call Anthropic Messages API via REST.

    Supports web_search tool when included in the tools list.
    """
    import json as _json
    import urllib.request
    import urllib.error

    url = "https://api.anthropic.com/v1/messages"
    data = _json.dumps(payload).encode("utf-8")
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            resp_body = resp.read().decode("utf-8", errors="replace")
            return _json.loads(resp_body)
    except urllib.error.HTTPError as e:
        err_txt = e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else str(e)
        logger.error(f"[error] Anthropic REST call failed: HTTP {e.code}: {err_txt}")
        return {"error": {"code": e.code, "message": err_txt}}
    except Exception as e:
        logger.error(f"[error] Anthropic REST call exception: {e}")
        return {"error": {"message": str(e)}}


def run_openai_web_search(model_key: str, query: str, json_structure: bool = True) -> Dict[str, Any]:
    """Run OpenAI GPT-5 family and deep research models with web search tool.
    
    Note: Deep research models (o4-mini-deep-research) are agentic and conduct 
    multi-step research, which can take 10+ minutes for complex queries.

    Args:
        model_key: One of 'openai-gpt-5', 'openai-gpt-5-mini', 'openai-gpt-5-nano', 
                   'openai-o4-mini-deep-research'.
        query: The search query text.
        json_structure: Whether to request structured JSON response or plain text.

    Returns:
        A result dictionary consistent with other runners, including usage and cost.
    """
    if not OPENAI_API_KEY:
        return {"model": model_key, "query": query, "error": "OPENAI_API_KEY not set"}

    logger.info(f"--- Running query for model: {model_key} (web search) ---")
    logger.info(f"Query: {query}\n")

    # Map model_key to OpenAI model id
    model_map = {
        "openai-gpt-5": "gpt-5",
        "openai-gpt-5-mini": "gpt-5-mini",
        "openai-gpt-5-nano": "gpt-5-nano",
        "openai-o4-mini-deep-research": "o4-mini-deep-research",
    }
    model_id = model_map.get(model_key, "gpt-5")

    prompt = get_universal_prompt(query, json_structure)

    try:
        started_at = time.perf_counter()
        
        # Use the OpenAI Python SDK's `responses.create` for the web_search_preview tool
        response = OPENAI_CLIENT.responses.create(
            model=model_id,
            tools=[{"type": "web_search_preview"}],
            input=prompt,
        )

        elapsed_s = time.perf_counter() - started_at

        # Extract output text and citations from the SDK response object
        # This structure is based on the previous REST implementation and may need validation
        message_text = response.output_text or ""
        # Remove thinking tags if present
        message_text = _remove_thinking_tags(message_text)
        
        # Usage tokens
        usage = response.usage
        prompt_tokens = usage.input_tokens if usage else 0
        completion_tokens = usage.output_tokens if usage else 0

        # Search tool calls count
        web_calls = 0
        if response.output:
            for item in response.output:
                if item.type == "web_search_call":
                    action = getattr(item, 'action', None) or getattr(getattr(item, 'web_search_call', {}), 'action', None)
                    if not action or action == "search":
                        web_calls += 1

        # Costs
        p = PRICING.get(model_key, {})
        token_cost = ((prompt_tokens / 1_000_000) * float(p.get("input", 0.0))) + (
            (completion_tokens / 1_000_000) * float(p.get("output", 0.0))
        )
        search_cost = float(p.get("search_per_request", 0.0)) * web_calls
        total_cost = token_cost + search_cost

        # Parse JSON and citations
        parsed_response = _parse_json_from_text(message_text)
        citations: List[Dict[str, str]] = []
        if response.output:
             for item in response.output:
                if item.type == "message":
                    for content_item in item.content:
                        if content_item.type == "output_text":
                            for annotation in getattr(content_item, 'annotations', []):
                                if annotation.type == "url_citation":
                                    url = getattr(annotation, 'url', None)
                                    title = getattr(annotation, 'title', "")
                                    if url:
                                        citations.append({"uri": url, "title": title})

        result = {
            "model": model_key,
            "query": query,
            "response": parsed_response if parsed_response is not None else message_text,
            "citations": citations,
            "usage": {
                "input_tokens": prompt_tokens,
                "output_tokens": completion_tokens,
                "total_tokens": prompt_tokens + completion_tokens,
                "estimated_token_cost_usd": f"${token_cost:.6f}",
                "estimated_search_cost_usd": f"${search_cost:.6f}",
                "estimated_total_cost_usd": f"${total_cost:.6f}",
                "web_search_calls": web_calls,
                "elapsed_seconds": round(elapsed_s, 3),
            },
        }
    except Exception as e:
        result = {"model": model_key, "query": query, "error": str(e)}

    return result



def _extract_text_from_rest_response(response_json: Dict[str, Any]) -> str:
    """Extract plain text from REST response JSON.

    Returns empty string if not found.
    """
    try:
        candidates = response_json.get("candidates", [])
        if not candidates:
            return ""
        parts = candidates[0].get("content", {}).get("parts", [])
        if parts and isinstance(parts[0], dict):
            return parts[0].get("text", "") or ""
    except Exception:
        pass
    return ""


def _extract_citations_from_rest_response(response_json: Dict[str, Any]) -> List[Dict[str, str]]:
    """Extract web citations from REST response JSON grounding metadata.

    Returns list of {uri, title} dictionaries.
    """
    citations: List[Dict[str, str]] = []
    try:
        candidates = response_json.get("candidates", [])
        if not candidates:
            return citations
        grounding_md = candidates[0].get("groundingMetadata", {})
        chunks = grounding_md.get("groundingChunks", [])
        for g in chunks:
            web = g.get("web", {}) if isinstance(g, dict) else {}
            uri = web.get("uri", "")
            title = web.get("title", "")
            if uri:
                citations.append({"uri": uri, "title": title})
    except Exception as e:
        logger.warning(f"[warn] Failed to parse REST citations: {e}")
    return citations


def _count_tokens_with_client(model_name: str, text: str) -> int:
    """Count tokens using the google-genai client.

    Falls back to a heuristic (len/4) if counting is unsupported.
    """
    try:
        resp = CLIENT.models.count_tokens(model=model_name, contents=text)
        # New SDK returns an object with total_tokens; handle dict fallback too
        total = getattr(resp, "total_tokens", None)
        if total is None and isinstance(resp, dict):
            total = resp.get("total_tokens") or resp.get("totalTokens")
        return int(total) if total is not None else max(1, len(text) // 4)
    except Exception as e:
        logger.warning(f"SDK token counting failed for model '{model_name}': {e}. Falling back to heuristic.")
        return max(1, len(text) // 4)


def _collect_flat_rows_for_export(all_results: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """Collect flattened rows from results for tabular export.

    Each row contains: model, query, item, price_huf, source, link, total_usd, elapsed_seconds.

    Args:
        all_results: List of result dictionaries from models.

    Returns:
        List of flat row dictionaries.
    """
    rows: List[Dict[str, str]] = []
    for res in all_results:
        model_name = str(res.get("model", ""))
        query = str(res.get("query", ""))
        usage = res.get("usage", {}) if isinstance(res, dict) else {}
        total_usd_str = ""
        try:
            total_usd_str = str(usage.get("estimated_total_cost_usd", "")).replace("$", "").strip()
        except Exception:
            total_usd_str = ""
        try:
            elapsed = str(usage.get("elapsed_seconds", "")).strip()
        except Exception:
            elapsed = ""

        response = res.get("response")
        table_rows = []
        plain_text_response = None
        
        if isinstance(response, dict):
            table_rows = response.get("table", [])
        elif isinstance(response, list):
            table_rows = response
        elif isinstance(response, str):
            # First, try to parse the string as JSON. If that fails, try parsing as Markdown.
            parsed_json = _parse_json_from_text(response)
            if isinstance(parsed_json, dict):
                table_rows = parsed_json.get("table", [])
            elif isinstance(parsed_json, list):
                table_rows = parsed_json
            else:
                markdown_table = _parse_markdown_table_from_text(response)
                if markdown_table:
                    table_rows = markdown_table
                else:
                    # This is a plain text response - store it as-is
                    plain_text_response = response

        # If we have structured table rows, export them
        if isinstance(table_rows, list) and len(table_rows) > 0:
            for row in table_rows:
                rows.append(
                    {
                        "model": model_name,
                        "query": query,
                        "item": str(row.get("item", "")),
                        "price_huf": str(row.get("price_huf", "")),
                        "source": str(row.get("source", "")),
                        "link": str(row.get("link", "")),
                        "total_usd": total_usd_str,
                        "elapsed_seconds": elapsed,
                    }
                )
        # If we have a plain text response, export it as a single row
        elif plain_text_response:
            rows.append(
                {
                    "model": model_name,
                    "query": query,
                    "response_text": plain_text_response,  # Full plain text response
                    "price_huf": "",
                    "source": "",
                    "link": "",
                    "total_usd": total_usd_str,
                    "elapsed_seconds": elapsed,
                    "is_plain_text": True,  # Flag to indicate this is a plain text response
                }
            )
    return rows


def _format_seconds_with_comma(seconds_value: Any) -> str:
    """Format elapsed seconds with decimal comma instead of decimal point.
    
    Args:
        seconds_value: The elapsed seconds value (can be string, float, or None).
        
    Returns:
        Formatted string with decimal comma (e.g., "1,234" instead of "1.234").
    """
    if not seconds_value:
        return ""
    
    try:
        # Convert to float if it's a string
        if isinstance(seconds_value, str):
            seconds_float = float(seconds_value)
        else:
            seconds_float = float(seconds_value)
        
        # Format with 3 decimal places and replace dot with comma
        return f"{seconds_float:.3f}".replace(".", ",")
    except Exception:
        return str(seconds_value)


def export_results_to_excel(file_path: str, all_results: List[Dict[str, Any]]) -> None:
    """Export combined results to an Excel .xlsx file.

    Creates a sheet named 'Results' with columns: model, item, price_huf, source, link.

    Args:
        file_path: Destination path for the Excel file.
        all_results: List of result dictionaries from different models.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        logger.warning(
            "[warn] openpyxl not available for Excel export. "
            "Install with: pip install openpyxl. Skipping .xlsx export.",
        )
        return

    rows = _collect_flat_rows_for_export(all_results)
    
    # Separate plain text responses from structured responses
    plain_text_rows = [r for r in rows if r.get("is_plain_text")]
    structured_rows = [r for r in rows if not r.get("is_plain_text")]
    
    # Keep only top-3 cheapest per model+query by price_huf for structured responses
    grouped: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for r in structured_rows:
        key = (r.get("model", ""), r.get("query", ""))
        grouped.setdefault(key, []).append(r)
    filtered_rows: List[Dict[str, str]] = []
    for key, items in grouped.items():
        def price_key(x: Dict[str, str]) -> float:
            try:
                return float(x.get("price_huf", "") or 0)
            except Exception:
                return float("inf")
        items_sorted = sorted(items, key=price_key)
        filtered_rows.extend(items_sorted[:3])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Write structured responses with full columns
    if filtered_rows:
        headers = ["model", "query", "item", "price_huf", "source", "link", "total_usd", "elapsed_seconds"]
        ws.append(headers)

        for r in filtered_rows:
            # Convert price to number when possible
            price_str = r.get("price_huf", "")
            try:
                price_val: Optional[float] = float(price_str) if str(price_str).strip() else None
            except Exception:
                price_val = None
            # Parse total cost
            total_str = r.get("total_usd", "")
            try:
                total_val: Optional[float] = float(total_str) if str(total_str).strip() else None
            except Exception:
                total_val = None
            row_idx = ws.max_row + 1
            ws.append([
                r.get("model", ""),
                r.get("query", ""),
                r.get("item", ""),
                price_val if price_val is not None else r.get("price_huf", ""),
                r.get("source", ""),
                None,
                total_val if total_val is not None else r.get("total_usd", ""),
                _format_seconds_with_comma(r.get("elapsed_seconds", "")),
            ])
            # Make link clickable using HYPERLINK formula to avoid DataValidation/pandas deps
            url = r.get("link", "") or ""
            if url:
                link_col = 6
                cell_ref = f"{get_column_letter(link_col)}{row_idx}"
                ws[cell_ref].value = f'=HYPERLINK("{url}", "{url}")'
    
    # Write plain text responses with minimal columns (no price_huf, source, link)
    if plain_text_rows:
        # Add separator if there were structured rows
        if filtered_rows:
            ws.append([])  # Empty row as separator
            ws.append(["=== Plain Text Responses ==="])
            ws.append([])
        
        headers_plain = ["model", "query", "response", "total_usd", "elapsed_seconds"]
        ws.append(headers_plain)
        
        for r in plain_text_rows:
            total_str = r.get("total_usd", "")
            try:
                total_val: Optional[float] = float(total_str) if str(total_str).strip() else None
            except Exception:
                total_val = None
            
            ws.append([
                r.get("model", ""),
                r.get("query", ""),
                r.get("response_text", ""),  # Full text, no truncation
                total_val if total_val is not None else r.get("total_usd", ""),
                _format_seconds_with_comma(r.get("elapsed_seconds", "")),
            ])

    try:
        wb.save(file_path)
        logger.info(f"Excel export written to {file_path}")
    except Exception as e:
        logger.error(f"[error] Failed to write Excel file '{file_path}': {e}")


# --- Main Execution ---

if __name__ == "__main__":
    # Dynamically build model runners from PRICING config
    PROVIDER_TO_RUNNER = {
        "google": run_gemini_query,
        "perplexity": run_perplexity_query,
        "xai": run_grok_live_search_query,
        "openai": run_openai_web_search,
        "anthropic": run_anthropic_web_search,
    }

    model_runners = {}
    for model_key, model_info in PRICING.items():
        provider = model_info.get("provider")
        runner_func = PROVIDER_TO_RUNNER.get(provider)
        if runner_func:
            # Normalize all runners to a simple `runner(query, json_structure)` signature
            model_runners[model_key] = runner_func

    # Timestamped filenames to avoid overwrite
    ts = datetime.now().strftime("%Y%m%d-%H%M")
    xlsx_path = f"search_results-{ts}.xlsx"

    # Select models to run
    # Build model list based on enabled flags and key presence
    MODELS_TO_TEST: List[str] = []
    for mk, meta in PRICING.items():
        if not meta.get("enabled", True):
            continue
        provider = meta.get("provider")
        if provider == "perplexity" and not PERPLEXITY_API_KEY:
            continue
        if provider == "xai" and not XAI_API_KEY:
            continue
        if provider == "openai" and not OPENAI_API_KEY:
            continue
        if provider == "anthropic" and not ANTHROPIC_API_KEY:
            continue
        if provider == "google" and not GEMINI_API_KEY:
            continue
        MODELS_TO_TEST.append(mk)

    all_results: List[Dict[str, Any]] = []
    for model_name in MODELS_TO_TEST:
        for query_config in QUERIES:
            query = query_config["query"]
            json_structure = query_config.get("json_structure", True)

            runner = model_runners.get(model_name)
            if not runner:
                logger.warning(f"[warn] No runner for model '{model_name}', skipping.")
                continue
            result = runner(model_name, query, json_structure)
            # measure elapsed time on the result by timing around the call if needed
            all_results.append(result)
            # reduce noisy logs; omit per-result dumps

    # Write Excel export (includes total cost per call)
    export_results_to_excel(xlsx_path, all_results)

    logger.info(f"All queries processed. Results saved to '{xlsx_path}'.")
